from geographiclib.geodesic import Geodesic
from LatLon23 import string2latlon, LatLon
from geomag import declination
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import config
import pynmea2
import subprocess
import os
import xlrd

def Vector2Polar(N,E):
    Mag = np.sqrt(N**2 + E**2)
    Az = 90 - np.arctan2(N, E) * 180/np.pi
    if Az <= 0:
        Az = Az + 360
    return(Az, Mag)

def bullcalculate(Lat, Long):
    if config.becoord is None:
        bull = ''
    else:
        if len(Lat)>0 and len(Long)>0:
            try:
                coord1 = config.becoord
                coord2 = string2latlon(Lat, Long, 'H% %d% %M')

                rbg = Geodesic.WGS84.Inverse(float(coord1.lat), float(coord1.lon), float(coord2.lat), float(coord2.lon))

                bulldist = round(rbg['s12']*0.000539957) # meters to NM
                bullaz = round(rbg['azi1'] - declination(float(coord1.lat), float(coord1.lon)))
                if bullaz < 0:
                    bullaz = 360 + bullaz
                bull = str(bullaz) + '/' + str(bulldist)
            except:
                bull = ''
        else:
                bull = ''
    return bull


def append_df_to_excel(filename, df, sheet_name, startrow=0,truncate_sheet=False,**to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """


    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    ws = writer.book[sheet_name]
    for column_cells in ws.columns:
        length = max(map(lambda cell: len(str(cell.value)) if cell.value else 0, column_cells))
        ws.column_dimensions[column_cells[0].column_letter].width = length+3
    # save the workbook
    writer.save()

def updatefillins(filename):
    try:
        # try to open an existing workbook
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = load_workbook(filename)

        benamerng = writer.book.defined_names['BEname']
        cells = []

        for title, coord in benamerng.destinations:
            ws = writer.book[title]
            ws[coord] = config.bename.upper()
        for title, coord in writer.book.defined_names['BELat'].destinations:
            ws = writer.book[title]
            ws[coord] = str(config.belat)
        for title, coord in writer.book.defined_names['BELong'].destinations:
            ws = writer.book[title]
            ws[coord] = str(config.belong)
        if len(config.csname)>0:
            for title, coord in writer.book.defined_names['cs'].destinations:
                ws = writer.book[title]
                ws[coord] = str(config.csname)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except:
        # file does not exist yet, we will create it
        print('Error writing Fill In Values')

    # save the workbook
    writer.save()

def to_gpsnmea(df,filename):
    df = df.rename(columns={"Time (UTC)": "TIME"})
    df['LAT'] = df['LAT'].str.replace(" ", '')
    df['LONG'] = df['LONG'].str.replace(" ", '')
    filename = filename.replace('Debrief Card','GPS Trail').replace('xlsx','gps')
    gpsfile = open(filename, 'a')
    for index, row in df.iterrows():
        mvar = float(row.THDG) - float(row.MHDG)
        if mvar < 0:
            mvarH = 'W'
        else:
            mvarH = 'E'
        GS = f'{row.GS:05.1f}'
        GTRK = f'{row.GTRK:05.1f}'
        MHDG = f'{row.MHDG:05.1f}'
        dtime = row.TIME.strftime('%H%M%S')
        ddate = row.TIME.strftime('%d%m%y')
        alt = str(round(float(row.ALT) * 0.3048))

        RMC = pynmea2.ZDA('GP', 'RMC', (dtime, 'A', row.LAT[1:], row.LAT[0], row.LONG[1:], row.LONG[0], GS, GTRK, ddate, f'{mvar:05.1f}', mvarH))
        GGA = pynmea2.GGA('GP', 'GGA', (dtime, row.LAT[1:], row.LAT[0], row.LONG[1:], row.LONG[0], '1', '', '', alt, 'M', '', '', '', '0'))
        VTG = pynmea2.GGA('GP', 'VTG', (GTRK, 'T', MHDG, 'M', f'{row.GS:06.2f}', 'N', '', 'K'))

        gpsfile.write(str(RMC) + "\n")
        gpsfile.write(str(GGA) + "\n")
        gpsfile.write(str(VTG) + "\n")
    gpsfile.close()
    #os.system('cmd .\output\GPSBabel\gpsbabel -i nmea -f ' + filename + ' -x interpolate,time=10 -o nmea -F "testtrack.gps"')
    #subprocess.Popen(r'explorer /select,'+ filename )
    os.startfile(filename, 'open')

def jassm_report_match(debrief_filename, jassm_report):

    jreport = pd.ExcelFile(jassm_report)
    wpngroups = []

    for sheet in jreport.sheet_names:
        if 'JASSMGRP' in sheet:
            #print(sheet)
            df = jreport.parse(sheet, skiprows=5, index_col=None, na_values=['NA'])
            df['wpngroup'] = str(sheet) + " " + 'MSN'+ df['Msn'].astype(str)
            wpngroups.append(df)
    if len(wpngroups)>0:
        wpns = pd.read_excel(debrief_filename, sheet_name='Combined',index_col=None, na_values=['NA'])
        wpns.astype({'TGT LAT':str,'TGT LONG':str,'TGT ELEV':str,'BULL':str}).dtypes
        wpns['TGT LAT'] = wpns['TGT LAT'].astype(str)
        wpns['TGT LONG'] = wpns['TGT LONG'].astype(str)
        wpns['TGT ELEV'] = wpns['TGT ELEV'].astype(str)
        wpns['BULL'] = wpns['BULL'].astype(str)

        for i, row in wpns.iterrows():
            for group in wpngroups:
                for j, rows in group.iterrows():
                    if wpns.at[i,'TGT Name'] == group.at[j,'wpngroup']:
                        wpns.at[i, 'TGT Name'] = group.at[j,'Mission Name']
                        wpns.at[i, 'TGT LAT'] = group.at[j, 'Tgt Latitude']
                        wpns.at[i, 'TGT LONG'] = group.at[j,'Tgt Longitude']
                        wpns.at[i, 'TGT ELEV'] = str(group.at[j,'Tgt Elev (HAE)']) + "' HAE"
                        try:
                            wpns.at[i, 'BULL'] = bullcalculate(wpns.at[i, 'TGT LAT'], wpns.at[i, 'TGT LONG'])
                        except:
                            wpns.at[i, 'BULL']  = ''
        wpns = wpns.fillna('')
        wpns = wpns.replace('nan', '', regex=True)

        return wpns
    else:
        return pd.DataFrame()

def jassm_report_match(debrief_filename='none', jassm_report='jrep.xlsm'):

    if debrief_filename == 'none':
        debrief_filename = QFileDialog.getOpenFileName(self, 'Open Debrief Card Excel File',
                                    os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop'),
                                    "Excel File (*.xlsx)")
    if jassm_report == 'none':
        jassm_report = QFileDialog.getOpenFileName(self, 'Open JMPS JASSM Report Excel File',
                                    os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop'),
                                    "Excel File (*.xlsx)")
    jreport = pd.ExcelFile(jassm_report)
    wpngroups = []

    for sheet in jreport.sheet_names:
        if 'JASSMGRP' in sheet:
            print(sheet)
            df = jreport.parse(sheet, skiprows=5, index_col=None)
            df['wpngroup'] = str(sheet) + " " + 'MSN'+ df['Msn'].astype(str)
            wpngroups.append(df)
    wpns = pd.read_excel(debrief_filename, sheet_name='Combined',index_col=None,na_filter= False)
    wpns.astype({'TGT LAT':str,'TGT LONG':str,'TGT ELEV':str,'BULL':str}).dtypes
    wpns['TGT LAT'] = wpns['TGT LAT'].astype(str)
    wpns['TGT LONG'] = wpns['TGT LONG'].astype(str)
    wpns['TGT ELEV'] = wpns['TGT ELEV'].astype(str)
    wpns['BULL'] = wpns['BULL'].astype(str)

    for i, row in wpns.iterrows():
        for group in wpngroups:
            for j, rows in group.iterrows():
                if wpns.at[i,'TGT Name'] == group.at[j,'wpngroup']:
                    wpns.at[i, 'TGT Name'] = group.at[j,'Mission Name']
                    wpns.at[i, 'TGT LAT'] = group.at[j, 'Tgt Latitude']
                    wpns.at[i, 'TGT LONG'] = group.at[j,'Tgt Longitude']
                    wpns.at[i, 'TGT ELEV'] = str(group.at[j,'Tgt Elev (HAE)']) + "' HAE"
                    try:
                        wpns.at[i, 'BULL'] = bullcalculate(wpns.at[i, 'TGT LAT'], wpns.at[i, 'TGT LONG'])
                    except:
                        wpns.at[i, 'BULL']  = ''
    wpns = wpns.fillna('')
    wpns = wpns.replace('nan', '', regex=True)
    return wpns